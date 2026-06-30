from __future__ import annotations

import argparse
import hashlib
import json
import re
from collections import Counter
from pathlib import Path
from typing import Any

import pdfplumber
from openpyxl import load_workbook
from pypdf import PdfReader

try:
    from tools.term_corrections import apply_correction, load_corrections
except ModuleNotFoundError:
    from term_corrections import apply_correction, load_corrections


CATEGORY_COUNTS = {
    "mechanical": 1227,
    "electrical": 970,
    "customer_review": 251,
    "meeting": 58,
    "business": 198,
}

TERM_CORRECTIONS_PATH = Path(__file__).resolve().parent / "data" / "term_corrections.json"

CATEGORY_LABELS = {
    "mechanical": "机械专业词汇",
    "electrical": "电气专业词汇",
    "customer_review": "客户评审/问题表达",
    "meeting": "会议常用口语",
    "business": "商务英语句子",
}

LIGATURES = str.maketrans({"ﬁ": "fi", "ﬂ": "fl", "’": "'", "“": '"', "”": '"'})
VARIANT_SPLIT = re.compile(r"[；;]|/(?=[A-Za-z])")
CJK = re.compile(r"[\u3400-\u9fff]")
MECHANICAL_ENGLISH_REPAIRS = {
    471: "C gun",
    472: "X gun (P gun)",
}
ELECTRICAL_ENGLISH_REPAIRS = {
    284: "C gun",
    285: "X gun (P gun)",
}
CUSTOMER_ENGLISH_REPAIRS = {
    27: "Separate the power cables from the communication cables.",
    33: (
        "If the part is skewed and the rack isn't put in place correctly, are we "
        "able to avoid like a potential crash there, or are we thinking that the "
        "cameras will be able to catch that?"
    ),
    56: "RFQ package: the customer's initial project release standard",
    79: (
        "Laser-welding fixture shielding gas: gas flows from the cylinder or "
        "workshop pipeline through the solenoid valve and pressure switch to "
        "the manifold, then to each gas-blowing port."
    ),
    95: "We need to be very gentle in the money",
    116: "If the payload is under 90, you need to use steel, especially for the pin.",
    129: "Could we jump to 90 station?",
    163: "MINO USA",
    172: "Make sure the clamp arm doesn't exceed the capacity of the clamp",
    188: "The tip dresser is 90 degrees respect to the robot.",
    220: (
        "If we have the robot riser at two meters, then we need the glue stand "
        "much higher, then we need a maintenance stand for the glue, maintenance "
        "stair for the robot...additional safety for maintenance stairs and so on. "
        "If we can manage to have one meter, the glue stand will be much lower, "
        "we have much less work to do."
    ),
    223: (
        "These are designed for like Amazon package shipment not industrial. "
        "You got those parts on there with sharp edges and all that stuff, I don't "
        "know how well it'll hold up, right? I don't know that we use belts like "
        "that. It seems like the belts we use have serrated pattern on them, so "
        "they can handle more wear and tear."
    ),
    229: "Customer",
    238: "Customer",
    248: "Proton",
    249: (
        "Gripper stand needs to be under 80 degrees in SS11 project and make it "
        "to be horizontal as much as possible."
    ),
    62: "Customer",
}
MEETING_ENGLISH_REPAIRS = {
    21: "Sorry?",
}
PHRASE_CATEGORIES = {"customer_review", "meeting", "business"}
LEGITIMATE_NUMERIC_PHRASE_ENDINGS = {
    ("customer_review", 244): "50X50",
}


def phrase_footer_issue(
    category: str,
    source_index: int,
    english: str,
) -> str | None:
    """Return a semantic footer-contamination reason for phrase records."""
    if category not in PHRASE_CATEGORIES:
        return None
    if (category, source_index) in LEGITIMATE_NUMERIC_PHRASE_ENDINGS:
        return None

    value = normalize_text(english)
    if re.search(r"\d+\s*$", value):
        if re.search(r"(?:\(\s*\)|[,?;:])(?:\s*(?:\(\s*\)|[,?;:]))*\s+\d+\s*$", value):
            return "trailing empty-parenthesis/punctuation footer"
        return "suspicious trailing page number"
    return None


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).translate(LIGATURES)
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r" *\n *", " ", text)
    return text.strip(" \t\r\n")


def clean_primary_english(english: str) -> str:
    value = normalize_text(english)
    first = VARIANT_SPLIT.split(value, maxsplit=1)[0].strip()
    first = re.sub(r"\s*\([^)]*[A-Z][A-Z0-9& ]*\)\s*$", "", first).strip()
    first = re.sub(r"\s*\([^)]*[\u3400-\u9fff][^)]*\)\s*$", "", first).strip()
    return first or value


def make_stable_id(category: str, source_index: int, english: str) -> str:
    payload = f"{category}|{source_index}|{normalize_text(english).casefold()}"
    digest = hashlib.sha256(payload.encode("utf-8")).hexdigest()[:16]
    return f"{category[:3]}_{source_index:04d}_{digest}"


def _find_sources(source_dir: Path) -> tuple[Path, Path]:
    spreadsheets = list(source_dir.glob("*.xlsx"))
    pdfs = list(source_dir.glob("*.pdf"))
    if len(spreadsheets) != 1 or len(pdfs) != 1:
        raise ValueError("assets must contain exactly one .xlsx and one .pdf source")
    return spreadsheets[0], pdfs[0]


def _extract_term_tables(pdf_path: Path) -> tuple[list[list[str]], list[list[str]]]:
    rows: list[list[str]] = []
    with pdfplumber.open(pdf_path) as document:
        for page in document.pages:
            for table in page.extract_tables():
                for row in table:
                    if (
                        row
                        and len(row) >= 4
                        and row[0]
                        and re.fullmatch(r"\d+", normalize_text(row[0]))
                    ):
                        rows.append(
                            [
                                normalize_text(row[0]),
                                normalize_text(row[1]),
                                normalize_text(row[2]),
                                normalize_text(row[3]),
                            ]
                        )

    chunks: list[list[list[str]]] = [[]]
    previous = 0
    for row in rows:
        number = int(row[0])
        if chunks[-1] and number <= previous:
            chunks.append([])
        chunks[-1].append(row)
        previous = number
    if [len(chunk) for chunk in chunks] != [1227, 970]:
        raise ValueError(f"unexpected PDF term table sizes: {[len(chunk) for chunk in chunks]}")
    return chunks[0], chunks[1]


def _xlsx_terms(workbook_path: Path, sheet_index: int, expected: int) -> dict[int, dict[str, str]]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[sheet_index]]
    terms: dict[int, dict[str, str]] = {}
    for row in sheet.iter_rows(values_only=True):
        if not isinstance(row[0], (int, float)):
            continue
        number = int(row[0])
        if not 1 <= number <= expected:
            continue
        terms[number] = {
            "english": normalize_text(row[1]),
            "chinese": normalize_text(row[2]),
            "note": normalize_text(row[3]),
        }
    workbook.close()
    if len(terms) != expected:
        raise ValueError(f"unexpected Excel term count in sheet {sheet_index}: {len(terms)}")
    return terms


def _trim_card_tail(text: str, known_headings: set[str]) -> str:
    lines = [normalize_text(line) for line in text.splitlines()]
    while lines and (
        not lines[-1]
        or re.fullmatch(r"\d{1,3}", lines[-1])
        or lines[-1] in known_headings
        or re.fullmatch(r"(规划|设计) *- *.+", lines[-1])
    ):
        lines.pop()
    return normalize_text("\n".join(lines))


def _extract_cards(section: str, known_headings: set[str]) -> list[dict[str, str]]:
    marker = re.compile(r"(?m)^#(\d+)([^\n]*)\n")
    matches = list(marker.finditer(section))
    cards: list[dict[str, str]] = []
    for index, match in enumerate(matches):
        body = section[
            match.end() : matches[index + 1].start() if index + 1 < len(matches) else len(section)
        ]
        if "英文：" not in body or "音标：" not in body:
            continue
        _, body = body.split("英文：", 1)
        english, remainder = body.split("\n音标：", 1)
        if "\n中文：" in remainder:
            phonetic, chinese = remainder.split("\n中文：", 1)
        else:
            phonetic, chinese = remainder, ""
        cards.append(
            {
                "number": str(match.group(1)),
                "section": normalize_text(match.group(2)),
                "english": normalize_text(english),
                "phonetic": normalize_text(phonetic),
                "chinese": _trim_card_tail(chinese, known_headings),
            }
        )
    return cards


def _extract_phrase_sections(pdf_path: Path) -> tuple[list[dict[str, str]], ...]:
    reader = PdfReader(str(pdf_path))
    text = "\n".join((page.extract_text() or "") for page in reader.pages[85:])
    customer_heading = "客户评审观点及问题 - 机械"
    meeting_heading = "会议常用口语"
    business_heading = "商务英语交流常用必背句子"
    starts = [text.find(customer_heading), text.find(meeting_heading), text.find(business_heading)]
    if any(start < 0 for start in starts) or starts != sorted(starts):
        raise ValueError(f"PDF phrase section headings not found in order: {starts}")

    customer_text = text[starts[0] : starts[1]]
    meeting_text = text[starts[1] : starts[2]]
    business_text = text[starts[2] :]

    customer_headings = {
        "机械设计",
        "仿真",
        "规划",
        "规划 - 工艺规划",
        "规划 - 布局",
        "其他",
        "T客户",
    }
    meeting_headings = {f"{index}.{label}" for index, label in enumerate(
        [
            "开头语",
            "介绍参会人",
            "设备权限",
            "会议目的",
            "讨论议题",
            "澄清问题",
            "表达观点",
            "同意与不同意",
            "提出建议",
            "确认理解",
            "总结",
            "行动项",
            "会议结束语",
        ],
        start=1,
    )}
    business_headings = {
        "英文邮件",
        "商务谈判",
        "会议沟通",
        "工作汇报",
        "电话沟通",
        "客户沟通",
        "项目管理",
        "团队协作",
        "问题解决",
        "其他常用表达",
    }

    customer = _extract_cards(customer_text, customer_headings)
    meeting = _extract_cards(meeting_text, meeting_headings)
    business = _extract_cards(business_text, business_headings)

    # The source intentionally counts this title as meeting item #1.
    if meeting and CJK.search(meeting[0]["english"]):
        meeting[0] = {
            "number": "1",
            "section": "概览",
            "english": "Common Meeting English Expressions",
            "phonetic": "/ˈkɑːmən ˈmiːtɪŋ ˈɪŋɡlɪʃ ɪkˈsprɛʃənz/",
            "chinese": "常用会议英语口语汇总",
        }
    else:
        meeting.insert(
            0,
            {
                "number": "1",
                "section": "概览",
                "english": "Common Meeting English Expressions",
                "phonetic": "/ˈkɑːmən ˈmiːtɪŋ ˈɪŋɡlɪʃ ɪkˈsprɛʃənz/",
                "chinese": "常用会议英语口语汇总",
            },
        )

    # Two source rows have blank translations; keep them explicit and auditable.
    for card in customer:
        if card["english"] == "Stellantis/Autodie" and not card["chinese"]:
            card["chinese"] = "斯特兰蒂斯 / Autodie"
    if len(customer) != 251 or len(meeting) != 58 or len(business) != 198:
        raise ValueError(
            "unexpected phrase counts: "
            f"customer={len(customer)}, meeting={len(meeting)}, business={len(business)}"
        )
    return customer, meeting, business


def _short_chinese(text: str) -> str:
    value = re.split(r"[（(；;。]", normalize_text(text), maxsplit=1)[0]
    return value or normalize_text(text)


def _example(category: str, english: str, chinese: str) -> tuple[str, str]:
    primary = clean_primary_english(english)
    short_zh = _short_chinese(chinese)
    if category == "mechanical":
        return (
            f"The technician checked the {primary} before the production trial.",
            f"技术员在生产试运行前检查了{short_zh}。",
        )
    if category == "electrical":
        return (
            f"The engineer verified the {primary} during system commissioning.",
            f"工程师在系统调试期间确认了{short_zh}。",
        )
    return normalize_text(english), normalize_text(chinese)


def _make_record(
    category: str,
    source_index: int,
    english: str,
    phonetic: str,
    chinese: str,
    note: str = "",
    section: str = "",
) -> dict[str, Any]:
    english = normalize_text(english)
    phonetic = normalize_text(phonetic)
    chinese = normalize_text(chinese)
    note = normalize_text(note)
    primary = clean_primary_english(english)
    record_id = make_stable_id(category, source_index, english)
    example_en, example_zh = _example(category, english, chinese)
    audio_text = re.sub(r"[；;]", ", ", english)
    return {
        "id": record_id,
        "category": category,
        "categoryLabel": CATEGORY_LABELS[category],
        "sourceIndex": source_index,
        "kind": "TERM" if category in {"mechanical", "electrical"} else "PHRASE",
        "section": normalize_text(section),
        "english": english,
        "primaryEnglish": primary,
        "phonetic": phonetic,
        "chinese": chinese,
        "note": note,
        "exampleEn": example_en,
        "exampleZh": example_zh,
        "audioText": normalize_text(audio_text),
        "audioAsset": f"audio/{record_id}.ogg",
    }


def build_content(source_dir: Path, output_path: Path, report_path: Path) -> list[dict[str, Any]]:
    workbook_path, pdf_path = _find_sources(source_dir)
    mechanical_pdf, electrical_pdf = _extract_term_tables(pdf_path)
    mechanical_xlsx = _xlsx_terms(workbook_path, 2, 1227)
    electrical_xlsx = _xlsx_terms(workbook_path, 3, 970)
    reviewed_corrections = load_corrections(TERM_CORRECTIONS_PATH)
    unmatched_corrections = set(reviewed_corrections)

    records: list[dict[str, Any]] = []
    for category, table, workbook_terms in (
        ("mechanical", mechanical_pdf, mechanical_xlsx),
        ("electrical", electrical_pdf, electrical_xlsx),
    ):
        for row in table:
            number = int(row[0])
            source = workbook_terms[number]
            english = source["english"] or row[1]
            chinese = source["chinese"] or row[3].replace("备注：机械设计", "").strip()
            phonetic = row[2]
            if category == "mechanical" and number == 129 and not chinese:
                chinese = "模型梁"
            if number in {1006, 751} and CJK.search(english) and "AGV" in chinese:
                english = "AGV - Automated Guided Vehicle"
                chinese = "无人搬运车"
                phonetic = "/ˌeɪ dʒiː ˈviː - ˈɔːtəmeɪtɪd ˈɡaɪdɪd ˈviːəkəl/"
            record = _make_record(
                category,
                number,
                english,
                phonetic,
                chinese,
                source["note"],
            )
            term_repairs = (
                MECHANICAL_ENGLISH_REPAIRS
                if category == "mechanical"
                else ELECTRICAL_ENGLISH_REPAIRS
            )
            if number in term_repairs:
                repaired_english = term_repairs[number]
                record.update(
                    {
                        "english": repaired_english,
                        "primaryEnglish": repaired_english,
                        "exampleEn": repaired_english,
                        "audioText": repaired_english,
                    }
                )
            correction_identity = (category, number)
            if correction := reviewed_corrections.get(correction_identity):
                record = apply_correction(record, correction)
                unmatched_corrections.remove(correction_identity)
            records.append(record)

    if unmatched_corrections:
        raise ValueError(
            f"reviewed corrections did not match content: {sorted(unmatched_corrections)}"
        )

    customer, meeting, business = _extract_phrase_sections(pdf_path)
    for category, cards in (
        ("customer_review", customer),
        ("meeting", meeting),
        ("business", business),
    ):
        for source_index, card in enumerate(cards, start=1):
            record = _make_record(
                category,
                source_index,
                card["english"],
                card["phonetic"] or "/fəˈnɛtɪk nɑːt əˈveɪləbəl/",
                card["chinese"],
                section=card["section"],
            )
            phrase_repairs = (
                CUSTOMER_ENGLISH_REPAIRS
                if category == "customer_review"
                else MEETING_ENGLISH_REPAIRS if category == "meeting" else {}
            )
            if source_index in phrase_repairs:
                repaired_english = phrase_repairs[source_index]
                record.update(
                    {
                        "english": repaired_english,
                        "primaryEnglish": repaired_english,
                        "phonetic": "/phonetic not available/",
                        "exampleEn": repaired_english,
                        "audioText": repaired_english,
                    }
                )
            records.append(record)

    phrase_footer_issues = [
        {
            "category": item["category"],
            "sourceIndex": item["sourceIndex"],
            "english": item["english"],
            "reason": reason,
        }
        for item in records
        if (
            reason := phrase_footer_issue(
                item["category"],
                item["sourceIndex"],
                item["english"],
            )
        )
    ]
    if phrase_footer_issues:
        raise ValueError(
            f"phrase footer contamination detected: {phrase_footer_issues[:20]}"
        )

    counts = Counter(item["category"] for item in records)
    if dict(counts) != CATEGORY_COUNTS:
        raise ValueError(f"unexpected category counts: {dict(counts)}")
    if len({item["id"] for item in records}) != len(records):
        raise ValueError("duplicate stable IDs detected")

    required = (
        "english",
        "primaryEnglish",
        "phonetic",
        "chinese",
        "exampleEn",
        "exampleZh",
        "audioText",
        "audioAsset",
    )
    missing = [
        (item["id"], field)
        for item in records
        for field in required
        if not str(item[field]).strip()
    ]
    if missing:
        raise ValueError(f"missing required content: {missing[:20]}")
    unpronounceable = [
        (item["id"], field)
        for item in records
        for field in ("audioText",)
        if not re.search(
            r"(?<![A-Za-z0-9])(?=[A-Za-z0-9]*[A-Za-z])[A-Za-z0-9]{2,}",
            str(item[field]),
        )
    ]
    if unpronounceable:
        raise ValueError(f"English pronunciation text missing: {unpronounceable[:20]}")

    payload = {
        "schemaVersion": 1,
        "contentVersion": "2026.06.29",
        "sourceFiles": [workbook_path.name, pdf_path.name],
        "words": records,
    }
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    report = {
        "total": len(records),
        "counts": dict(counts),
        "uniqueIds": len({item["id"] for item in records}),
        "reviewedTermCorrections": [
            {
                "category": correction.category,
                "sourceIndex": correction.source_index,
                "originalEnglish": correction.original_english,
                "correctedEnglish": correction.corrected_english,
                "reason": correction.reason,
                "stableIdPreserved": True,
                "evidence": [
                    {
                        "organization": evidence.organization,
                        "url": evidence.url,
                        "accessedOn": evidence.accessed_on,
                    }
                    for evidence in correction.evidence
                ],
            }
            for correction in reviewed_corrections.values()
        ],
        "phraseFooterAudit": {
            "count": len(phrase_footer_issues),
            "issues": phrase_footer_issues,
            "allowlisted": [
                {
                    "category": category,
                    "sourceIndex": source_index,
                    "legitimateEnding": ending,
                }
                for (category, source_index), ending
                in sorted(LEGITIMATE_NUMERIC_PHRASE_ENDINGS.items())
            ],
        },
        "repairs": [
            {"category": "mechanical", "sourceIndex": 129, "field": "chinese", "value": "模型梁"},
            *[
                {
                    "category": category,
                    "sourceIndex": source_index,
                    "field": "english/primaryEnglish/exampleEn/audioText",
                    "value": english,
                    "stableIdPreserved": True,
                }
                for category, repairs in (
                    ("mechanical", MECHANICAL_ENGLISH_REPAIRS),
                    ("electrical", ELECTRICAL_ENGLISH_REPAIRS),
                )
                for source_index, english in repairs.items()
            ],
            {
                "category": "customer_review",
                "source": "Stellantis/Autodie",
                "field": "chinese",
                "value": "斯特兰蒂斯 / Autodie",
            },
            {
                "category": "meeting",
                "sourceIndex": 1,
                "field": "english/phonetic",
                "value": "normalized meeting title",
            },
            *[
                {
                    "category": "customer_review",
                    "sourceIndex": source_index,
                    "field": "english/primaryEnglish/exampleEn/audioText",
                    "value": english,
                    "stableIdPreserved": True,
                }
                for source_index, english in CUSTOMER_ENGLISH_REPAIRS.items()
            ],
            *[
                {
                    "category": "meeting",
                    "sourceIndex": source_index,
                    "field": "english/primaryEnglish/exampleEn/audioText",
                    "value": english,
                    "stableIdPreserved": True,
                }
                for source_index, english in MEETING_ENGLISH_REPAIRS.items()
            ],
        ],
    }
    report_path.parent.mkdir(parents=True, exist_ok=True)
    report_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    return records


def main() -> None:
    parser = argparse.ArgumentParser(description="Build MIearn auditable content JSON")
    parser.add_argument("--source-dir", type=Path, default=Path("assets"))
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("app/src/main/assets/content/words_v1.json"),
    )
    parser.add_argument(
        "--report",
        type=Path,
        default=Path("output/reports/content_report.json"),
    )
    args = parser.parse_args()
    records = build_content(args.source_dir, args.output, args.report)
    print(f"Built {len(records)} MIearn records -> {args.output}")


if __name__ == "__main__":
    main()
